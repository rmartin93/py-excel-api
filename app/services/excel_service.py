"""
Excel utility service for report generation.
Provides reusable functions for Excel operations, but does NOT contain template-specific logic.
Each template should have its own generator function for maximum flexibility.
"""

import io
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from app.core.config import get_settings
from app.core.logging import get_logger
from app.services.template_service import template_service
from app.models.reports import ReportRequest, ReportResponse

logger = get_logger(__name__)
settings = get_settings()


class ExcelUtilities:
    """Utility functions for Excel operations - NOT template-specific."""

    def __init__(self):
        logger.info("Excel utilities initialized")

    def create_workbook(self, sheet_name: str = "Report") -> openpyxl.Workbook:
        """Create a new Excel workbook with specified sheet name."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        return workbook

    def add_headers(self, worksheet, headers: List[str], row: int = 1, start_col: int = 1) -> None:
        """Add header row to worksheet."""
        for col_idx, header_name in enumerate(headers, start_col):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.value = header_name

    def populate_data_rows(self, worksheet, data_rows: List[Dict[str, Any]],
                          columns: List[str], start_row: int = 2, start_col: int = 1) -> int:
        """
        Populate data rows in worksheet.

        Returns:
            int: The next available row after data
        """
        current_row = start_row

        for row_data in data_rows:
            for col_idx, column_name in enumerate(columns, start_col):
                if column_name in row_data:
                    cell = worksheet.cell(row=current_row, column=col_idx)
                    cell.value = row_data[column_name]
            current_row += 1

        return current_row

    def add_calculated_totals_row(self, worksheet, columns: List[str], data_rows: List[Dict[str, Any]],
                                 totals_row: int, start_col: int = 1) -> None:
        """Add manually calculated totals row (for templates without SUM formulas)."""
        if not data_rows:
            return

        totals = {}

        # Calculate totals for numeric columns
        for row_data in data_rows:
            for column_name, value in row_data.items():
                if isinstance(value, (int, float)) and not column_name.lower().endswith('id'):
                    if column_name not in totals:
                        totals[column_name] = 0.0
                    totals[column_name] += value

        # Set totals in worksheet
        for col_idx, column_name in enumerate(columns, start_col):
            cell = worksheet.cell(row=totals_row, column=col_idx)

            if col_idx == start_col:  # First column gets "Total" label
                cell.value = "Total"
                cell.font = Font(bold=True)
            elif column_name in totals:
                cell.value = totals[column_name]
                cell.font = Font(bold=True)

    def add_sum_formulas(self, worksheet, columns: List[str], data_start_row: int,
                        data_end_row: int, totals_row: int, start_col: int = 1) -> None:
        """Add SUM formulas for numeric columns."""
        for col_idx, column_name in enumerate(columns, start_col):
            cell = worksheet.cell(row=totals_row, column=col_idx)

            if col_idx == start_col:  # First column gets "Total" label
                cell.value = "Total"
                cell.font = Font(bold=True)
            else:
                # Add SUM formula for data range
                column_letter = get_column_letter(col_idx)
                formula = f"=SUM({column_letter}{data_start_row}:{column_letter}{data_end_row})"
                cell.value = formula
                cell.font = Font(bold=True)

    def apply_header_formatting(self, worksheet, headers: List[str], header_row: int = 1,
                               start_col: int = 1, bg_color: str = "366092",
                               font_color: str = "FFFFFF") -> None:
        """Apply professional formatting to header row."""
        header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        header_font = Font(color=font_color, bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx in range(start_col, start_col + len(headers)):
            header_cell = worksheet.cell(row=header_row, column=col_idx)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment

    def apply_data_formatting(self, worksheet, columns: List[str], data_start_row: int,
                             data_end_row: int, start_col: int = 1) -> None:
        """Apply formatting to data rows."""
        data_alignment = Alignment(horizontal="left", vertical="center")

        for row_idx in range(data_start_row, data_end_row):
            for col_idx in range(start_col, start_col + len(columns)):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.alignment = data_alignment

    def apply_totals_formatting(self, worksheet, columns: List[str], totals_row: int,
                               start_col: int = 1, bg_color: str = "E7E6E6") -> None:
        """Apply formatting to totals row."""
        totals_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")

        for col_idx in range(start_col, start_col + len(columns)):
            totals_cell = worksheet.cell(row=totals_row, column=col_idx)
            totals_cell.fill = totals_fill
            totals_cell.font = Font(bold=True)

    def apply_borders(self, worksheet, start_row: int, end_row: int,
                     start_col: int, end_col: int) -> None:
        """Apply borders to a range of cells."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

    def auto_size_columns(self, worksheet, columns: List[str], start_col: int = 1,
                         default_width: int = 15) -> None:
        """Auto-size column widths."""
        for col_idx in range(start_col, start_col + len(columns)):
            column_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[column_letter].width = default_width

    def save_to_buffer(self, workbook: openpyxl.Workbook) -> io.BytesIO:
        """Save workbook to BytesIO buffer."""
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        workbook.close()
        return buffer

    def create_excel_table(self, worksheet, table_name: str, start_cell: str,
                          end_cell: str) -> None:
        """Create an Excel Table object in the worksheet."""
        table = Table(displayName=table_name, ref=f"{start_cell}:{end_cell}")
        # Add table to worksheet
        worksheet.add_table(table)


class ExcelService:
    """Main Excel service that routes to template-specific generators."""

    def __init__(self):
        self.utils = ExcelUtilities()
        logger.info("Excel service initialized with utilities")

    def generate_report(self, request: ReportRequest) -> ReportResponse:
        """
        Route to appropriate template-specific generator.
        Each template gets its own custom generation logic.
        """
        try:
            # Validate template exists
            validation = template_service.validate_template(request.template_name)
            if not validation["valid"]:
                return ReportResponse(
                    success=False,
                    data={},
                    message=validation["error"],
                    file_data=None
                )

            # Route to template-specific generator
            template_name = request.template_name.lower()

            if template_name == "template-1.xlsx":
                return self._generate_template_1_report(request)
            else:
                # Default/fallback generator for unknown templates
                return self._generate_generic_report(request)

        except Exception as e:
            logger.error(f"Error generating report: {e}")
            return ReportResponse(
                success=False,
                data={},
                message=f"Error generating report: {str(e)}",
                file_data=None
            )

    def _generate_template_1_report(self, request: ReportRequest) -> ReportResponse:
        """
        Template-specific generator for Template-1.xlsx
        This can be as custom/weird as needed for this specific template.
        """
        try:
            # Get template structure
            templates_response = template_service.list_templates()
            template_info = None
            for template in templates_response.data["templates"]:
                if template.filename == request.template_name:
                    template_info = template
                    break

            if not template_info:
                return ReportResponse(
                    success=False,
                    data={},
                    message=f"Template {request.template_name} not found",
                    file_data=None
                )

            # CREATE CUSTOM LOGIC FOR TEMPLATE-1 HERE
            workbook = self.utils.create_workbook("Financial Report")
            worksheet = workbook.active

            # Your specific Template-1 columns
            columns = template_info.columns
            rows_data = request.data.get("rows", [])

            # Custom layout for Template-1
            self.utils.add_headers(worksheet, columns, row=1)
            self.utils.populate_data_rows(worksheet, rows_data, columns, start_row=2)

            # For Template-1: Use calculated totals (since it has SUM formulas)
            # You can switch this to add_sum_formulas if your template uses Excel formulas
            totals_row = len(rows_data) + 2
            self.utils.add_calculated_totals_row(worksheet, columns, rows_data, totals_row)

            # Apply Template-1 specific formatting (always formatted for Template-1)
            self.utils.apply_header_formatting(worksheet, columns)
            self.utils.apply_data_formatting(worksheet, columns, 2, totals_row)
            self.utils.apply_totals_formatting(worksheet, columns, totals_row)
            self.utils.apply_borders(worksheet, 1, totals_row, 1, len(columns))
            self.utils.auto_size_columns(worksheet, columns)

            # Generate file
            excel_buffer = self.utils.save_to_buffer(workbook)

            # Create response
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"template_1_report_{timestamp}.xlsx"

            return ReportResponse(
                success=True,
                data={
                    "filename": output_filename,
                    "size": len(excel_buffer.getvalue()),
                    "template_used": request.template_name,
                    "generated_at": datetime.now().isoformat(),
                    "rows_processed": len(rows_data)
                },
                message="Template-1 report generated successfully",
                file_data=excel_buffer.getvalue()
            )

        except Exception as e:
            logger.error(f"Error generating Template-1 report: {e}")
            return ReportResponse(
                success=False,
                data={},
                message=f"Error generating Template-1 report: {str(e)}",
                file_data=None
            )

    def _generate_generic_report(self, request: ReportRequest) -> ReportResponse:
        """
        Fallback generator for unknown templates.
        Keep this simple - complex templates should get their own generator.
        """
        try:
            # Very basic generation for unknown templates
            workbook = self.utils.create_workbook("Report")
            worksheet = workbook.active

            # Try to extract some basic info
            rows_data = request.data.get("rows", [])
            if not rows_data:
                return ReportResponse(
                    success=False,
                    data={},
                    message="No data provided for report generation",
                    file_data=None
                )

            # Use first row keys as columns
            columns = list(rows_data[0].keys()) if rows_data else []

            # Basic layout
            self.utils.add_headers(worksheet, columns)
            next_row = self.utils.populate_data_rows(worksheet, rows_data, columns, start_row=2)

            # Basic formatting
            self.utils.apply_header_formatting(worksheet, columns)
            self.utils.auto_size_columns(worksheet, columns)

            excel_buffer = self.utils.save_to_buffer(workbook)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"generic_report_{timestamp}.xlsx"

            return ReportResponse(
                success=True,
                data={
                    "filename": output_filename,
                    "size": len(excel_buffer.getvalue()),
                    "template_used": request.template_name,
                    "generated_at": datetime.now().isoformat(),
                    "rows_processed": len(rows_data)
                },
                message="Generic report generated successfully",
                file_data=excel_buffer.getvalue()
            )

        except Exception as e:
            logger.error(f"Error generating generic report: {e}")
            return ReportResponse(
                success=False,
                data={},
                message=f"Error generating generic report: {str(e)}",
                file_data=None
            )


# Global instance
excel_service = ExcelService()
