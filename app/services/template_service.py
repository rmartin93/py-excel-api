"""
Template management service for Excel API.
Handles template discovery, validation, and metadata extraction.
"""

import os
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.worksheet.table import Table

from app.core.config import get_settings
from app.core.logging import get_logger
from app.models.templates import TemplateInfo, TemplateListResponse

logger = get_logger(__name__)
settings = get_settings()


class TemplateService:
    """Service for managing Excel templates."""

    def __init__(self):
        self.templates_dir = Path(settings.templates_dir)
        self.templates_dir.mkdir(exist_ok=True)
        logger.info(f"Template service initialized with directory: {self.templates_dir}")

    def list_templates(self) -> TemplateListResponse:
        """
        List all available Excel templates with metadata.

        Returns:
            TemplateListResponse: List of available templates
        """
        try:
            templates = []

            # Find all Excel files in templates directory
            excel_files = list(self.templates_dir.glob("*.xlsx")) + list(self.templates_dir.glob("*.xls"))

            for file_path in excel_files:
                try:
                    template_info = self._get_template_info(file_path)
                    templates.append(template_info)
                except Exception as e:
                    logger.warning(f"Could not process template {file_path.name}: {e}")
                    continue

            logger.info(f"Found {len(templates)} valid templates")

            return TemplateListResponse(
                success=True,
                data={"templates": templates},
                message=f"Found {len(templates)} templates"
            )

        except Exception as e:
            logger.error(f"Error listing templates: {e}")
            return TemplateListResponse(
                success=False,
                data={"templates": []},
                message=f"Error listing templates: {str(e)}"
            )

    def _get_template_info(self, file_path: Path) -> TemplateInfo:
        """
        Extract metadata from an Excel template file.

        Args:
            file_path: Path to the Excel file

        Returns:
            TemplateInfo: Template metadata
        """
        stat = file_path.stat()

        # Try to analyze the Excel file structure
        table_info = self._analyze_excel_structure(file_path)

        return TemplateInfo(
            name=file_path.stem.replace("_", " ").replace("-", " ").title(),
            filename=file_path.name,
            size=stat.st_size,
            last_modified=datetime.fromtimestamp(stat.st_mtime),
            description=f"Excel template with {table_info.get('tables', 0)} table(s)",
            columns=table_info.get('columns', []),
            sample_data=table_info.get('sample_data', {})
        )

    def _analyze_excel_structure(self, file_path: Path) -> Dict[str, Any]:
        """
        Analyze Excel file to extract table structure and sample data needs.

        Args:
            file_path: Path to Excel file

        Returns:
            Dict containing structure information
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            worksheet = workbook.active

            # Look for Excel tables
            tables = worksheet.tables
            columns = []
            sample_data = {}

            if tables:
                # Get columns from first table
                table_name = list(tables.keys())[0]
                table = tables[table_name]

                # Extract headers from table range
                min_col, min_row, max_col, _ = openpyxl.utils.range_boundaries(table.ref)

                for col in range(min_col, max_col + 1):
                    header_cell = worksheet.cell(row=min_row, column=col)
                    if header_cell.value:
                        column_name = str(header_cell.value).strip()
                        columns.append(column_name)

                        # Generate sample data based on column name
                        sample_data[column_name] = self._generate_sample_value(column_name)

                logger.info(f"Found table '{table_name}' with columns: {columns}")

            else:
                # No tables found, try to detect headers in first row
                for col in range(1, 20):  # Check first 20 columns
                    header_cell = worksheet.cell(row=1, column=col)
                    if header_cell.value:
                        column_name = str(header_cell.value).strip()
                        columns.append(column_name)
                        sample_data[column_name] = self._generate_sample_value(column_name)
                    else:
                        break

                logger.info(f"Found header row with columns: {columns}")

            workbook.close()

            return {
                "tables": len(tables),
                "columns": columns,
                "sample_data": sample_data
            }

        except Exception as e:
            logger.warning(f"Could not analyze Excel structure for {file_path.name}: {e}")
            return {"tables": 0, "columns": [], "sample_data": {}}

    def _generate_sample_value(self, column_name: str) -> Any:
        """
        Generate appropriate sample data based on column name.

        Args:
            column_name: Name of the column

        Returns:
            Sample value appropriate for the column type
        """
        column_lower = column_name.lower()

        # ID columns
        if 'id' in column_lower:
            return "RULE001"

        # Group/category columns (before amount check)
        if any(keyword in column_lower for keyword in ['group', 'center', 'category', 'type']):
            return "Sample Group"

        # Financial/numeric columns
        if any(keyword in column_lower for keyword in ['amount', 'rate', 'cost', 'price', 'total']):
            return 1000.50

        # Difference/variance columns
        if any(keyword in column_lower for keyword in ['diff', 'variance', 'delta']):
            return 0.0

        # Default to string
        return "Sample Value"

    def get_template_path(self, filename: str) -> Optional[Path]:
        """
        Get the full path to a template file.

        Args:
            filename: Name of the template file

        Returns:
            Path to template file if it exists, None otherwise
        """
        template_path = self.templates_dir / filename

        if template_path.exists() and template_path.suffix.lower() in ['.xlsx', '.xls']:
            return template_path

        return None

    def validate_template(self, filename: str) -> Dict[str, Any]:
        """
        Validate that a template file exists and is accessible.

        Args:
            filename: Name of the template file

        Returns:
            Dict with validation results
        """
        template_path = self.get_template_path(filename)

        if not template_path:
            return {
                "valid": False,
                "error": f"Template '{filename}' not found"
            }

        try:
            # Try to open the file
            workbook = openpyxl.load_workbook(template_path)
            workbook.close()

            return {
                "valid": True,
                "path": str(template_path),
                "size": template_path.stat().st_size
            }

        except Exception as e:
            return {
                "valid": False,
                "error": f"Cannot open template '{filename}': {str(e)}"
            }


# Global instance
template_service = TemplateService()
